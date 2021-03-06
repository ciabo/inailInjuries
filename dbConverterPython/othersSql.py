import openpyxl;
from pandas import *
import datetime;

#codici icd10

document = openpyxl.load_workbook('data/codici-ICD10-descrizioni.xlsx');
sheets = document.get_sheet_names();
list = [];
for i in range(0, len(sheets)):
    sheet = document.get_sheet_by_name(sheets[i]);
    if (i == 0):
        for j in range(2, 1581):
            tmp= [document.get_sheet_by_name(sheets[i]).cell(row=j,column=1).value,
                  document.get_sheet_by_name(sheets[i]).cell(row=j,column=4).value];
            list.append(tmp)
    elif (i==1):
        for j in range(2, 373):
            tmp = [document.get_sheet_by_name(sheets[i]).cell(row=j, column=1).value,
                   document.get_sheet_by_name(sheets[i]).cell(row=j, column=4).value];
            list.append(tmp);
    elif (i==2):
        for j in range(2, 8467):
            code = document.get_sheet_by_name(sheets[i]).cell(row=j, column=1).value +"."+ str(document.get_sheet_by_name(
                sheets[i]).cell(row=j, column=2).value);
            tmp = [code, document.get_sheet_by_name(sheets[i]).cell(row=j, column=3).value];
            list.append(tmp);
    elif (i==3):
        for j in range(2, 3319):
            code = document.get_sheet_by_name(sheets[i]).cell(row=j, column=1).value + "." + str(document.get_sheet_by_name(
                sheets[i]).cell(row=j, column=2).value);
            tmp = [code, document.get_sheet_by_name(sheets[i]).cell(row=j, column=3).value];
            list.append(tmp);

f = open('data/sql/SQLCodici.sql','w')
f.write("CREATE TABLE IF NOT EXISTS codici (id varchar(10) NOT NULL, descrizione varchar(100),PRIMARY KEY (id));\n")
for i in range(0,len(list)):
    f.write("INSERT INTO codici (id, descrizione) VALUES (\""+list[i][0]+"\",\""+list[i][1]+"\");\n");
f.write("INSERT INTO codici (id, descrizione) VALUES (\"997\",\"Other complications of procedures not elsewhere classified\");\n");
f.write("INSERT INTO codici (id, descrizione) VALUES (\"998\",\"Complications affecting specified body system not elsewhere classified\");\n");
f.write("CREATE VIEW tumori AS SELECT id, descrizione FROM codici WHERE descrizione LIKE \"%tumore maligno%\" or descrizione LIKE \"%tumori maligni%\" OR descrizione LIKE \"%carcinoma%\" OR descrizione LIKE \"%melanoma%\" OR descrizione LIKE \"%linfom%\" OR descrizione LIKE \"%Mesotelioma%\" OR descrizione LIKE \"%sarcoma%\" OR descrizione LIKE \"%leucemi%\";");
f.close()


#Nazioni

document = openpyxl.load_workbook('data/ucm.xlsx');
sheet = document.get_sheet_names();
list = [];
for i in range(3,297):
    tmp = [document.get_sheet_by_name(sheet[0]).cell(row=i, column=3).value,
           document.get_sheet_by_name(sheet[0]).cell(row=i, column=2).value];
    list.append(tmp);
    f = open('data/sql/SQLNazioni.sql', 'w')
    f.write("CREATE TABLE IF NOT EXISTS nazioni (id varchar(10) NOT NULL, nazione varchar(45),PRIMARY KEY (id));\n")
    for i in range(0, len(list)):
        if(list[i][1]=="ITALIA"):
            f.write("INSERT INTO nazioni (id, nazione) VALUES (\"ITAL\",\"ITALIA\");\n");
        else:
            f.write("INSERT INTO nazioni (id, nazione) VALUES (\"" + list[i][0] + "\",\"" + list[i][1] + "\");\n");
    f.close()


#Sedi inail

document = openpyxl.load_workbook('data/SediINAIL.xlsx');
sheet = document.get_sheet_names();
list = [];

province = openpyxl.load_workbook('data/provinceCoordinate.xlsx');
sheetProvince = province.get_sheet_names();

with open('data/province.txt') as f:
    content = f.readlines()
listaProvince=[];
for row in content:
    tmp=[]
    rowElements=row.split(",");
    for i in range(0,len(rowElements)):
        if(i!=2): #per evitare di copiare anche la regione
            tmp.append(rowElements[i]);
    listaProvince.append(tmp);

for i in range(2,2352):
    provincia="";
    siglaProvincia="";
    cap = document.get_sheet_by_name(sheet[0]).cell(row=i, column=10).value;
    if ((cap>=86010 and cap<=86049) or (cap==86100)):
        provincia="Campobasso";
        siglaProvincia="CB";
    elif((cap>=86070 and cap<=86097) or (cap==86170)):
        provincia = "Isernia";
        siglaProvincia = "IS";
    if ((cap>=33010 and cap<=33059) or (cap==33100)):
        provincia="Udine";
        siglaProvincia="UD";
    elif((cap>= 33070 and cap<=33099) or (cap==33170)):
        provincia = "Pordenone";
        siglaProvincia = "PN";
    if((cap>=8010 and cap<=8049) or cap==8100):
        provincia = "Nuoro";
        siglaProvincia = "NU";
    elif ((cap >= 9121 and cap <= 9134) or (cap >= 9010 and cap <= 9048)):
        provincia = "Cagliari";
        siglaProvincia = "CA";
    elif (cap >= 9010 and cap <= 9066):
        provincia = "Sud Sardegna";
        siglaProvincia = "SU";
    elif ((cap >= 9070 and cap <= 9099) or (str(cap).zfill(5)[:3]=="080") or (cap==9170)):
        provincia = "Oristano";
        siglaProvincia = "OR";
    if ((cap>=34010 and cap<=34018) or (cap>=34121 and cap<=34151)):
        provincia="Trieste";
        siglaProvincia="TS";
    elif((cap>=34070  and cap<=34079) or (cap==34170)):
        provincia = "Gorizia";
        siglaProvincia = "GO";
    else:
        cap=str(cap).zfill(5)[:3];
        for p in listaProvince:
            dim=len(p);
            if(dim==3):
                print(p)
            elif(dim==4):
                if (cap == p[2][:3] or cap==p[3][:3]):
                    provincia = p[0];
                    siglaProvincia = p[1];
                    break;
            else:
                if (cap == p[2][:3] or cap==p[3][:3] or cap==p[4][:3]):
                    provincia = p[0];
                    siglaProvincia = p[1];
                    break;
    latitudineProvincia=0;
    longitudineProvincia=0;
    for j in range(1,7983):
        if (provincia==province.get_sheet_by_name(sheetProvince[0]).cell(row=j,column=1).value):
            latitudineProvincia=province.get_sheet_by_name(sheetProvince[0]).cell(row=j,column=21).value;
            longitudineProvincia=province.get_sheet_by_name(sheetProvince[0]).cell(row=j,column=22).value;
    tmp = [document.get_sheet_by_name(sheet[0]).cell(row=i, column=3).value,  # id - provincia
           document.get_sheet_by_name(sheet[0]).cell(row=i, column=2).value,  # sede - denominazione sede
           document.get_sheet_by_name(sheet[0]).cell(row=i, column=10).value,  # cap - cap servito
           provincia,siglaProvincia,latitudineProvincia,longitudineProvincia];
    list.append(tmp);

    f = open('data/sql/SQLSedi.sql', 'w')
    f.write("CREATE TABLE IF NOT EXISTS sedi (id INTEGER NOT NULL AUTO_INCREMENT, codice INTEGER, sede varchar(45), cap INTEGER,"
            "provincia varchar(30),siglaProvincia varchar(2),"
            "latitudineProvincia VARCHAR(40), longitudineProvincia VARCHAR(40), PRIMARY KEY (id));\n")
    for i in range(0, len(list)):
        f.write("INSERT INTO sedi (codice, sede, cap,provincia,siglaProvincia,latitudineProvincia,longitudineProvincia) VALUES (" + str(list[i][0]) + ",\"" + list[i][1] + "\"," + str(list[i][2]).zfill(5) + ",\"" + str(list[i][3]) + "\",\"" + str(list[i][4]) + "\",\"" + str(list[i][5]) + "\",\"" + str(list[i][6]) + "\");\n");
    f.write("CREATE VIEW singolesedi AS SELECT DISTINCT codice, latitudineProvincia, longitudineProvincia from sedi;")
    f.close();



