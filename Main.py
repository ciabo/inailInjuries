import openpyxl;
from pandas import *
import datetime;

document = openpyxl.load_workbook('data/codici-ICD10-descrizioni.xlsx');

sheets = document.get_sheet_names();
list = [];
for i in range(0, len(sheets)):
    sheet = document.get_sheet_by_name(sheets[i]);
    if (i == 0):
        for j in range(2, 1581):
            tmp= [document.get_sheet_by_name(sheets[i]).cell(row=j,column=1).value,
                  document.get_sheet_by_name(sheets[i]).cell(row=j,column=4).value];
            list.append(tmp)
    elif (i==1):
        for j in range(2, 373):
            tmp = [document.get_sheet_by_name(sheets[i]).cell(row=j, column=1).value,
                   document.get_sheet_by_name(sheets[i]).cell(row=j, column=4).value];
            list.append(tmp);
    elif (i==2):
        for j in range(2, 8467):
            code = document.get_sheet_by_name(sheets[i]).cell(row=j, column=1).value +"."+ str(document.get_sheet_by_name(
                sheets[i]).cell(row=j, column=2).value);
            tmp = [code, document.get_sheet_by_name(sheets[i]).cell(row=j, column=3).value];
            list.append(tmp);
    elif (i==3):
        for j in range(2, 3319):
            code = document.get_sheet_by_name(sheets[i]).cell(row=j, column=1).value + "." + str(document.get_sheet_by_name(
                sheets[i]).cell(row=j, column=2).value);
            tmp = [code, document.get_sheet_by_name(sheets[i]).cell(row=j, column=3).value];
            list.append(tmp);

f = open('data/SQLCodici.sql','w')
f.write("CREATE TABLE IF NOT EXISTS codici (id varchar(10) NOT NULL, descrizione varchar(45),PRIMARY KEY (id));\n")
for i in range(0,len(list)):
    f.write("INSERT INTO codici (id, descrizione) VALUES (\""+list[i][0]+"\",\""+list[i][1]+"\");\n");
f.write("INSERT INTO codici (id, descrizione) VALUES (\"997\",\"Other complications of procedures not elsewhere classified\");\n");
f.write("INSERT INTO codici (id, descrizione) VALUES (\"998\",\"Complications affecting specified body system not elsewhere classified\");\n");
f.close()

document = openpyxl.load_workbook('data/ucm.xlsx');
sheet = document.get_sheet_names();
list = [];
for i in range(3,297):
    tmp = [document.get_sheet_by_name(sheet[0]).cell(row=i, column=3).value,
           document.get_sheet_by_name(sheet[0]).cell(row=i, column=2).value];
    list.append(tmp);
    f = open('data/SQLNazioni.sql', 'w')
    f.write("CREATE TABLE IF NOT EXISTS nazioni (id varchar(10) NOT NULL, nazione varchar(45),PRIMARY KEY (id));\n")
    for i in range(0, len(list)):
        if(list[i][1]=="ITALIA"):
            f.write("INSERT INTO nazioni (id, nazione) VALUES (\"ITAL\",\"ITALIA\");\n");
        else:
            f.write("INSERT INTO nazioni (id, nazione) VALUES (\"" + list[i][0] + "\",\"" + list[i][1] + "\");\n");
    f.close()

document = openpyxl.load_workbook('data/SediINAIL.xlsx');
sheet = document.get_sheet_names();
list = [];
for i in range(2,2350):
    tmp = [document.get_sheet_by_name(sheet[0]).cell(row=i, column=3).value, #id - provincia
           document.get_sheet_by_name(sheet[0]).cell(row=i, column=2).value, #sede - denominazione sede
           document.get_sheet_by_name(sheet[0]).cell(row=i, column=8).value, #cap - cap
           document.get_sheet_by_name(sheet[0]).cell(row=i, column=11).value, #email - email
           document.get_sheet_by_name(sheet[0]).cell(row=i, column=13).value];  #telefono - telefono
    list.append(tmp);
    f = open('data/SQLSedi.sql', 'w')
    f.write("CREATE TABLE IF NOT EXISTS sedi (id INTEGER NOT NULL AUTO_INCREMENT, codice INTEGER, sede varchar(45), cap INTEGER,"
            " email varchar(40), telefono varchar(40),PRIMARY KEY (id));\n")
    for i in range(0, len(list)):
        f.write("INSERT INTO sedi (codice, sede, cap, email, telefono) VALUES (" + str(list[i][0]) + ",\"" + list[i][1] + "\"," + str(list[i][2]) + ",\"" + list[i][3] + "\",\"" + str(list[i][4]) + "\");\n");
    f.close();

regioni = ["Abruzzo", "Basilicata", "Calabria", "Campania", "EmiliaRomagna", "FriuliVeneziaGiulia", "Lazio", "Liguria",
           "Lombardia", "Marche", "Molise", "Piemonte", "Puglia", "Sardegna", "Sicilia", "Toscana", "TrentinoAltoAdige",
           "Umbria", "ValledAosta", "Veneto"];
formatStr = '%d/%m/%Y';
f = open('data/SQLSemestrali.sql', 'w')
f.write("CREATE TABLE IF NOT EXISTS datisemestrali (id INTEGER NOT NULL AUTO_INCREMENT,regione VARCHAR(40), data DATE ,"
        "dataMorte DATE , sedeInail INTEGER, sesso TINYINT, luogoNascita VARCHAR(10), ICD10denunciato VARCHAR(10),"
        "ICD10accertato VARCHAR(10), asbestoCorrelata TINYINT, giorniIndennizzati INTEGER, settore VARCHAR(20),PRIMARY KEY (id));\n")
for regione in regioni:
    df = read_csv('data/dati/'+regione+'.csv')
    for i, row in enumerate(df.values):
        tmp=row.tolist();
        tmpsplitted=tmp[0].split(";")
        data = tmpsplitted[0]
        dataMorte= tmpsplitted[3]
        sedeInail=tmpsplitted[4]
        sesso=tmpsplitted[6]
        luogoNascita=tmpsplitted[7]
        ICD10denunciato=tmpsplitted[10]
        ICD10accertato=tmpsplitted[11]
        giorniIndennizzati=tmpsplitted[22]

        head, sep, tail = str(datetime.datetime.strptime(data, formatStr)).partition(' ');
        data=head;
        if (dataMorte!=""):
            head, sep, tail = str(datetime.datetime.strptime(dataMorte, formatStr)).partition(' ');
            dataMorte=head;
        if(tmpsplitted[3]=="M"):
            sesso=1;
        else:
            sesso=0;

        if (tmpsplitted[14] == "S"):
            asbestoCorrelata = 1;
        else:
            asbestoCorrelata = 0;

        if (tmpsplitted[23]=="I"):
            settore = "Industria";
        elif(tmpsplitted[23]=="A"):
            settore = "Agricoltura";
        elif(tmpsplitted[23]=="S"):
            settore= "Statale";
        else:
            settore= "Medico-Domestico";

        if(dataMorte!=""):
            f.write("INSERT INTO datisemestrali (regione, data, dataMorte, sedeInail, sesso, luogoNascita, ICD10denunciato, "
                    "ICD10accertato, asbestoCorrelata, giorniIndennizzati, settore) VALUES (\"" + regione + "\",\"" +
                    data + "\",\"" + dataMorte + "\"," +
                    str(sedeInail) + "," + str(sesso) + ",\"" + luogoNascita + "\",\"" + ICD10denunciato + "\",\"" +
                    ICD10accertato + "\", "+str(asbestoCorrelata)+", "+ str(giorniIndennizzati) +",\"" + settore + "\");\n");
        else:
            f.write("INSERT INTO datisemestrali (regione, data, sedeInail, sesso, luogoNascita, ICD10denunciato, "
                    "ICD10accertato, asbestoCorrelata, giorniIndennizzati, settore) VALUES (\"" + regione + "\",\"" +
                    data + "\"," +
                    str(sedeInail) + "," + str(sesso) + ",\"" + luogoNascita + "\",\"" + ICD10denunciato + "\",\"" +
                    ICD10accertato + "\", " + str(asbestoCorrelata) + ", " + str(
                giorniIndennizzati) + ",\"" + settore + "\");\n");
f.close();
